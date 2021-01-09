# -*- coding: utf-8 -*-
# @Author   :   YaMeng
# @File :   lesson_7.py
# @Software :   PyCharm
# @Time :   2021/1/8 20:02
# @company  :   湖南省零檬信息技术有限公司

# 读取了excel之后，要怎么去使用里面的数据呢？  -- return!!

# 写入数据
# 循环写入测试结果，也就是意味着要重复调用写入的功能。  --- 封装函数
import openpyxl
import requests
# def write_reslt(filename, sheetname,row,column,final_result):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb[sheetname]
#     sheet.cell(row=row, column=column).value = final_result  # 写入结果到excel
#     wb.save(filename) # 保存excel


'''
接口自动化的步骤：
1、需要通过excle准备好测试用例，使用代码自动读取excle里的测试用例 -- read_data
2、发送接口请求，得到响应结果 -- func
3、对比执行结果和预期结果 -- done
4、得到的最终结果最后写入到excle -- write_reslt
'''

def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row  # 获取到sheet里面的最大行数
    testcases = []  # 定义一个空列表，用来接收测试用例
    for i in range(2, max_row+1):  # range() 取头不取尾
        dict1 = dict(
        id = sheet.cell(row=i, column=1).value, # 取用例编号
        url = sheet.cell(row=i, column=5).value, # 取出接口地址url
        data = sheet.cell(row=i, column=6).value, # 取出请求参数body
        expect = sheet.cell(row=i, column=7).value) # 取出预期结果
        testcases.append(dict1) # 把字典追加到列表，存放所有的测试用例数据
    return testcases

def api_func(url,data):
    header = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    res = requests.post(url=url, json=data, headers=header)
    res_log = res.json()
    return res_log

def write_reslt(filename, sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result  # 写入结果到excel
    wb.save(filename) # 保存excel


# 接口自动化：
def execute_func(filename, sheetname):
    cases = read_data(filename,sheetname)
    for case in cases:
        case_id = case['id'] # 取出id字段
        # case_id = case.get('id')
        case_url = case['url'] # 取出url字段
        case_data = case['data'] # 取出data字段
        case_expect = case['expect'] # 取出expect字段
        case_data = eval(case_data)  # eval()作用：把被字符串包裹着的python表达式，引号去掉。取里面的值
        case_expect = eval(case_expect) # 转换格式
        case_expect_msg = case_expect['msg'] # 取出预期结果里面的msg具体数据
        real_result = api_func(url=case_url, data=case_data)   # 调用发送请求的函数 -- 传入参数
        real_result_msg = real_result['msg'] # 取出实际结果里面的msg具体数据
        print('预期结果为:{}'.format(case_expect_msg))
        print('实际结果为:{}'.format(real_result_msg))
        if case_expect_msg == real_result_msg:
            print('这条测试用例通过！')
            final_res = 'pass'
        else:
            print('这条用例不通过！！！！！！！')
            final_res = 'fail'
        print('*' * 30)
        write_reslt(filename,sheetname,case_id + 1,8,final_res)

execute_func('test_case_api.xlsx','register')
execute_func('test_case_api.xlsx','login')

# 写了3个函数外加一个判断
# 读取数据函数
# 发送请求函数
# 做了预期结果和实际结果的判断
# 回写测试结果到excel